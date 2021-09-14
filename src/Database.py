from openpyxl import load_workbook

class Database:
    def __init__(self, path):
        self.workbook = load_workbook(path)
        self.sheet = self.workbook['info']

    def get_header(self):
        header = []
        for row in self.sheet.iter_rows(min_row=1, max_row=1):
            for item in row:
                header.append(str(item.value))
        return header
    
    def get_index_header(self, name):
        return self.get_header().index(name) + 1

    def get_index_username(self, username):
        return self.get_record_col('Username').index(username) + 2

    def get_record_col(self, col_name):
        records = []
        col = self.get_index_header(col_name)
        for col in self.sheet.iter_cols(min_row=2, min_col=col, max_col=col):
            for item in col:
                records.append(str(item.value))
        return records

    def exist_username(self, username):
        for item in self.get_record_col('Username'):
            if username == str(item):
                return True
        return False

    def get_password(self, username):
        if self.exist_username(username):
            return self.sheet.cell(row = self.get_index_username(username) , column = self.get_index_header('Password')).value

    def get_name(self, username):
        if self.exist_username(username):
            return self.sheet.cell(row = self.get_index_username(username) , column = self.get_index_header('Name')).value
